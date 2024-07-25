from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter


def append_result(sheet, item, expected_results, test_date, pass_fail, tested_by, comments, font):
    last_row = sheet.max_row + 1
    sheet[f"A{last_row}"] = item
    sheet[f"B{last_row}"] = expected_results
    sheet[f"C{last_row}"] = test_date
    sheet[f"D{last_row}"] = pass_fail
    sheet[f"E{last_row}"] = tested_by
    sheet[f"F{last_row}"] = comments

    sheet.row_dimensions[last_row].height = 25

    green_fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")
    red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")

    if pass_fail == "Y":
        sheet[f"D{last_row}"].fill = green_fill
    elif pass_fail == "N":
        sheet[f"D{last_row}"].fill = red_fill

    for cell in sheet[last_row]:
        cell.font = font
        cell.alignment = Alignment(horizontal="center",vertical="top")
        # cell.alignment = Alignment(horizontal="center", vertical="top")


def setup_workbook():
    try:
        workbook = load_workbook(filename="UAT - Release 10-Amvac.xlsx")
    except FileNotFoundError:
        workbook = Workbook()
        sheet = workbook.active
        titles = ["Item", "Expected Results", "Test Date", "Pass(Y/N)", "Tested by", "Comments"]
        sheet.append(titles)

    sheet = workbook.active
    calibri_font = Font(name='Arial', size=11)
    bold_font = Font(name='Arial', size=11, bold=True)

    for col_num in range(1, len(titles) + 1):
        cell = sheet[f"{get_column_letter(col_num)}1"]
        cell.font = bold_font
        cell.alignment = Alignment(horizontal="center", vertical="top")

    for row in sheet.iter_rows(min_row=2, min_col=1, max_col=sheet.max_column, max_row=sheet.max_row):
        for cell in row:
            cell.font = calibri_font
            cell.alignment = Alignment(vertical="top")

    sheet.row_dimensions[1].height = 30
    column_widths = {'A': 40, 'B': 50, 'C': 40, 'D': 35, 'E': 35, 'F': 60}
    for column, width in column_widths.items():
        sheet.column_dimensions[column].width = width

    return workbook, sheet


def save_workbook(workbook, filename="UAT - Release 10-Auto.xlsx"):
    workbook.save(filename)
